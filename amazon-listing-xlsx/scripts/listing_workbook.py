#!/usr/bin/env python3
import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


SOURCE_SHEET = "Product Info"
TARGET_SHEET = "listings"

REQUIRED_SOURCE_HEADERS = [
    "Item Code",
    "Product Name",
    "Main Color",
    "Main Material",
    "Description",
]

REQUIRED_SOURCE_FEATURE_HEADERS = [f"Product Features {index}" for index in range(1, 6)]

TARGET_HEADERS = [
    "itemCode",
    "productTitle",
    "productDescription",
    "searchTerms",
    "bulletPoint1",
    "bulletPoint2",
    "bulletPoint3",
    "bulletPoint4",
    "bulletPoint5",
]


class WorkbookError(Exception):
    pass


def cell_to_json_value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def load_workbook_sheet(path: Path, sheet_name: str) -> Worksheet:
    if not path.exists():
        raise WorkbookError(f"Input file does not exist: {path}")
    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:
        raise WorkbookError(f"Input file is not a readable .xlsx workbook: {path}") from exc
    if sheet_name not in workbook.sheetnames:
        raise WorkbookError(f"Workbook does not contain required sheet: {sheet_name}")
    return workbook[sheet_name]


def read_headers(sheet: Worksheet) -> list[str]:
    headers: list[str] = []
    for cell in sheet[1]:
        value = "" if cell.value is None else str(cell.value).strip()
        headers.append(value)
    return headers


def validate_source_headers(headers: list[str]) -> None:
    missing = [header for header in REQUIRED_SOURCE_HEADERS if header not in headers]
    missing.extend(header for header in REQUIRED_SOURCE_FEATURE_HEADERS if header not in headers)
    if missing:
        raise WorkbookError("Missing required headers: " + ", ".join(missing))


def source_rows(path: Path) -> list[dict[str, Any]]:
    sheet = load_workbook_sheet(path, SOURCE_SHEET)
    headers = read_headers(sheet)
    validate_source_headers(headers)

    rows: list[dict[str, Any]] = []
    for row_index in range(2, sheet.max_row + 1):
        values = [cell_to_json_value(sheet.cell(row_index, column).value) for column in range(1, len(headers) + 1)]
        source = dict(zip(headers, values))
        if not any(value != "" for value in source.values()):
            continue

        item_code = str(source.get("Item Code", "")).strip()
        if not item_code:
            raise WorkbookError(f"Source row {row_index} is missing Item Code")

        rows.append(
            {
                "rowIndex": row_index,
                "itemCode": item_code,
                "source": source,
            }
        )

    if not rows:
        raise WorkbookError("No product data rows found")
    return rows


def read_generated_rows(path: Path) -> list[dict[str, Any]]:
    try:
        data = json.loads(path.read_text())
    except Exception as exc:
        raise WorkbookError(f"Generated listing JSON is not readable: {path}") from exc

    if not isinstance(data, list):
        raise WorkbookError("Generated listing JSON must be an array")

    for index, row in enumerate(data):
        if not isinstance(row, dict):
            raise WorkbookError(f"Generated row at index {index} must be an object")

    return data


def normalize_generated_row(row: dict[str, Any]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for header in TARGET_HEADERS:
        value = row.get(header, "")
        text = "" if value is None else str(value)
        normalized[header] = text.replace("\r", " ").replace("\n", " ").strip()
    return normalized


def validate_generated_rows(source: list[dict[str, Any]], generated: list[dict[str, Any]]) -> None:
    if len(generated) != len(source):
        raise WorkbookError(f"Generated row count {len(generated)} does not match source row count {len(source)}")

    source_by_row = {row["rowIndex"]: row for row in source}
    seen_row_indexes: set[Any] = set()
    for row in generated:
        row_index = row.get("rowIndex")
        if row_index is None:
            raise WorkbookError("Generated row is missing rowIndex")
        if row_index in seen_row_indexes:
            raise WorkbookError(f"Duplicate generated rowIndex: {row_index}")
        seen_row_indexes.add(row_index)
        if row_index not in source_by_row:
            raise WorkbookError(f"Generated row has unknown rowIndex: {row_index}")

        expected_item_code = source_by_row[row_index]["itemCode"]
        actual_item_code = str(row.get("itemCode", "")).strip()
        if actual_item_code and actual_item_code != expected_item_code:
            raise WorkbookError(
                f"Generated itemCode {actual_item_code} does not match source itemCode {expected_item_code}"
            )

    missing_row_indexes = [row["rowIndex"] for row in source if row["rowIndex"] not in seen_row_indexes]
    if missing_row_indexes:
        raise WorkbookError(f"Missing generated rowIndex: {missing_row_indexes[0]}")


def write_output_workbook(source_path: Path, generated_json_path: Path, output_path: Path) -> None:
    source = source_rows(source_path)
    generated = read_generated_rows(generated_json_path)
    validate_generated_rows(source, generated)

    generated_by_row = {row["rowIndex"]: row for row in generated}
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = TARGET_SHEET
    sheet.append(TARGET_HEADERS)

    for source_row in source:
        generated_row = dict(generated_by_row[source_row["rowIndex"]])
        generated_row["itemCode"] = source_row["itemCode"]
        normalized = normalize_generated_row(generated_row)
        sheet.append([normalized[header] for header in TARGET_HEADERS])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_output_path = output_path.with_name(f".{output_path.stem}.tmp{output_path.suffix}")
    try:
        workbook.save(temp_output_path)
        validate_output_workbook(source_path, temp_output_path)
        temp_output_path.replace(output_path)
    except Exception:
        if temp_output_path.exists():
            temp_output_path.unlink()
        raise


def validate_output_workbook(source_path: Path, output_path: Path) -> None:
    source = source_rows(source_path)
    if not output_path.exists():
        raise WorkbookError(f"Output file does not exist: {output_path}")
    try:
        workbook = load_workbook(output_path, data_only=True)
    except Exception as exc:
        raise WorkbookError(f"Output workbook is not readable: {output_path}") from exc
    if workbook.sheetnames != [TARGET_SHEET]:
        raise WorkbookError(
            f"Only sheet allowed is {TARGET_SHEET}; found sheets: {workbook.sheetnames}"
        )
    sheet = workbook[TARGET_SHEET]
    headers = read_headers(sheet)
    if headers != TARGET_HEADERS:
        raise WorkbookError(f"Output headers do not match target schema: {headers}")

    for row_index in range(2, sheet.max_row + 1):
        values = {header: sheet.cell(row_index, column).value for column, header in enumerate(TARGET_HEADERS, start=1)}
        missing = [header for header, value in values.items() if value in (None, "")]
        if missing:
            raise WorkbookError(f"Output row {row_index} has empty required fields: {', '.join(missing)}")

    output_count = max(sheet.max_row - 1, 0)
    if output_count != len(source):
        raise WorkbookError(f"Output row count {output_count} does not match source row count {len(source)}")


def default_output_path(source_path: Path) -> Path:
    return source_path.with_name(f"{source_path.stem}.amazon-listings.xlsx")


def command_export_source(args: argparse.Namespace) -> None:
    rows = source_rows(Path(args.source))
    payload = json.dumps(rows, ensure_ascii=False, indent=2)
    if args.output:
        output = Path(args.output)
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(payload)
    else:
        print(payload)


def command_write_output(args: argparse.Namespace) -> None:
    source_path = Path(args.source)
    output_path = Path(args.output) if args.output else default_output_path(source_path)
    write_output_workbook(source_path, Path(args.generated_json), output_path)
    print(f"Wrote Amazon listing workbook: {output_path}")


def command_validate_output(args: argparse.Namespace) -> None:
    validate_output_workbook(Path(args.source), Path(args.output))
    print("Output workbook is valid")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Prepare and validate Amazon listing XLSX workbooks.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    export_parser = subparsers.add_parser("export-source", help="Export Product Info rows as JSON for listing generation.")
    export_parser.add_argument("source")
    export_parser.add_argument("--output")
    export_parser.set_defaults(func=command_export_source)

    write_parser = subparsers.add_parser("write-output", help="Write generated listing JSON into target XLSX schema.")
    write_parser.add_argument("source")
    write_parser.add_argument("generated_json")
    write_parser.add_argument("--output")
    write_parser.set_defaults(func=command_write_output)

    validate_parser = subparsers.add_parser("validate-output", help="Validate generated XLSX against source workbook.")
    validate_parser.add_argument("source")
    validate_parser.add_argument("output")
    validate_parser.set_defaults(func=command_validate_output)
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        args.func(args)
    except WorkbookError as exc:
        print(str(exc), file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
