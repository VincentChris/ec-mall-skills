import json
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "amazon-listing-xlsx" / "scripts" / "listing_workbook.py"


SOURCE_HEADERS = [
    "Item Code",
    "Product Name",
    "Main Color",
    "Main Material",
    "Assembled Length(inch)",
    "Assembled Width(inch)",
    "Assembled Height(inch)",
    "Product weight(pound)",
    "Package Size-Length (inch)",
    "Package Size-Width (inch)",
    "Package Size-Height (inch)",
    "Package Size-Weight (pound)",
    "Description",
    "Product Features 1",
    "Product Features 2",
    "Product Features 3",
    "Product Features 4",
    "Product Features 5",
    "Product Features 6",
    "Product Features 7",
    "Product Features 8",
    "Product Features 9",
    "Product Features 10",
    "Product Main Image",
    "Notes",
]


TARGET_HEADERS = [
    "itemCode",
    "productTitle",
    "productDescription",
    "searchTerms",
    "bulletPoint1",
    "bulletPoint2",
    "bulletPoint3",
    "bulletPoint4",
    "bulletPoint5",
]


def create_source_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Info"
    ws.append(SOURCE_HEADERS)
    ws.append([
        "SKU-1",
        "Expandable Hardside Luggage",
        "Black",
        "ABS",
        20,
        12,
        28,
        22,
        30,
        18,
        12,
        24,
        "Durable suitcase with TSA lock and spinner wheels.",
        "Expandable design",
        "TSA lock",
        "Spinner wheels",
        "Nested storage",
        "Fully lined interior",
        "",
        "",
        "",
        "",
        "",
        "https://example.com/main.jpg",
        "",
    ])
    ws.append([
        "SKU-2",
        "Vintage Luggage Set",
        "Green",
        "ABS",
        "",
        "",
        "",
        "",
        27,
        17,
        11,
        12,
        "Suitcase with duffel bag and toiletry bag.",
        "Duffel integration",
        "Expandable capacity",
        "Double spinner wheels",
        "Side mounted TSA lock",
        "Silicone handles",
        "",
        "",
        "",
        "",
        "",
        "https://example.com/main2.jpg",
        "",
    ])
    wb.save(path)


def run_cli(*args: str, cwd: Path = ROOT) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, str(SCRIPT), *args],
        cwd=cwd,
        text=True,
        capture_output=True,
        check=False,
    )


def create_output_workbook(path: Path, headers: list[str], row: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "listings"
    ws.append(headers)
    ws.append(row)
    wb.save(path)


def create_output_workbook_with_extra_sheet(path: Path, headers: list[str], row: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "listings"
    ws.append(headers)
    ws.append(row)
    extra = wb.create_sheet("extra")
    extra["A1"] = "noise"
    wb.save(path)


def test_export_source_writes_row_json(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "rows.json"
    create_source_workbook(source)

    result = run_cli("export-source", str(source), "--output", str(output))

    assert result.returncode == 0, result.stderr
    rows = json.loads(output.read_text())
    assert len(rows) == 2
    assert rows[0]["rowIndex"] == 2
    assert rows[0]["itemCode"] == "SKU-1"
    assert rows[0]["source"]["Product Name"] == "Expandable Hardside Luggage"
    assert rows[1]["rowIndex"] == 3
    assert rows[1]["itemCode"] == "SKU-2"


def test_export_source_fails_when_required_header_missing(tmp_path: Path) -> None:
    source = tmp_path / "bad-source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Info"
    ws.append(["Item Code", "Product Name"])
    ws.append(["SKU-1", "Bad row"])
    wb.save(source)

    result = run_cli("export-source", str(source))

    assert result.returncode == 2
    assert "Missing required headers" in result.stderr
    assert "Description" in result.stderr


def test_write_output_creates_target_schema(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    generated = tmp_path / "generated.json"
    output = tmp_path / "amazon-listings.xlsx"
    create_source_workbook(source)
    generated.write_text(json.dumps([
        {
            "rowIndex": 2,
            "itemCode": "SKU-1",
            "productTitle": "Expandable Hardside Luggage With TSA Lock Spinner Wheels Black",
            "productDescription": "<b>Travel Ready</b><br>Durable suitcase for organized trips.",
            "searchTerms": "carryon baggage spinner suitcase travel case",
            "bulletPoint1": "1. EXPANDABLE STORAGE - Adds flexible packing room.",
            "bulletPoint2": "2. TSA SECURITY - Helps protect packed belongings.",
            "bulletPoint3": "3. SMOOTH MOBILITY - Spinner wheels move easily.",
            "bulletPoint4": "4. ORGANIZED INTERIOR - Lining helps separate items.",
            "bulletPoint5": "5. EASY STORAGE - Nested shape saves closet space.",
        },
        {
            "rowIndex": 3,
            "itemCode": "SKU-2",
            "productTitle": "Vintage Luggage Set With Duffel Toiletry Bag Spinner Wheels Green",
            "productDescription": "<b>Complete Set</b><br>Includes suitcase, duffel, and toiletry bag.",
            "searchTerms": "baggage trunk roller holiday organizer",
            "bulletPoint1": "1. COMPLETE SET - Includes coordinated travel pieces.",
            "bulletPoint2": "2. DUFFEL SLEEVE - Attaches to suitcase handle.",
            "bulletPoint3": "3. SECURE LOCK - Side lock supports travel security.",
            "bulletPoint4": "4. EASY ROLLING - Double wheels support movement.",
            "bulletPoint5": "5. ABS SHELL - Hardside body supports frequent trips.",
        },
    ]))

    result = run_cli("write-output", str(source), str(generated), "--output", str(output))

    assert result.returncode == 0, result.stderr
    wb = load_workbook(output)
    assert wb.sheetnames == ["listings"]
    ws = wb["listings"]
    assert [cell.value for cell in ws[1]] == TARGET_HEADERS
    assert ws.max_row == 3
    assert ws["A2"].value == "SKU-1"
    assert ws["C2"].value == "<b>Travel Ready</b><br>Durable suitcase for organized trips."
    assert ws["D2"].value == "carryon baggage spinner suitcase travel case"
    assert ws["E2"].value == "1. EXPANDABLE STORAGE - Adds flexible packing room."
    assert ws["F2"].value == "2. TSA SECURITY - Helps protect packed belongings."
    assert ws["G2"].value == "3. SMOOTH MOBILITY - Spinner wheels move easily."
    assert ws["H2"].value == "4. ORGANIZED INTERIOR - Lining helps separate items."
    assert ws["I2"].value == "5. EASY STORAGE - Nested shape saves closet space."
    assert ws["B3"].value.startswith("Vintage Luggage Set")


def test_write_output_rejects_row_count_mismatch(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    generated = tmp_path / "generated.json"
    create_source_workbook(source)
    generated.write_text("[]")

    result = run_cli("write-output", str(source), str(generated))

    assert result.returncode == 2
    assert "Generated row count 0 does not match source row count 2" in result.stderr


def test_validate_output_accepts_matching_workbook(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    generated = tmp_path / "generated.json"
    output = tmp_path / "amazon-listings.xlsx"
    create_source_workbook(source)
    generated.write_text(json.dumps([
        {
            "rowIndex": 2,
            "itemCode": "SKU-1",
            "productTitle": "Title One",
            "productDescription": "<b>Description One</b>",
            "searchTerms": "alpha beta",
            "bulletPoint1": "1. ONE - Content",
            "bulletPoint2": "2. TWO - Content",
            "bulletPoint3": "3. THREE - Content",
            "bulletPoint4": "4. FOUR - Content",
            "bulletPoint5": "5. FIVE - Content",
        },
        {
            "rowIndex": 3,
            "itemCode": "SKU-2",
            "productTitle": "Title Two",
            "productDescription": "<b>Description Two</b>",
            "searchTerms": "gamma delta",
            "bulletPoint1": "1. ONE - Content",
            "bulletPoint2": "2. TWO - Content",
            "bulletPoint3": "3. THREE - Content",
            "bulletPoint4": "4. FOUR - Content",
            "bulletPoint5": "5. FIVE - Content",
        },
    ]))
    write_result = run_cli("write-output", str(source), str(generated), "--output", str(output))
    assert write_result.returncode == 0, write_result.stderr

    result = run_cli("validate-output", str(source), str(output))

    assert result.returncode == 0, result.stderr
    assert "Output workbook is valid" in result.stdout


def test_validate_output_rejects_wrong_header_order(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "bad-output.xlsx"
    create_source_workbook(source)
    create_output_workbook(
        output,
        [
            "itemCode",
            "productTitle",
            "searchTerms",
            "productDescription",
            "bulletPoint1",
            "bulletPoint2",
            "bulletPoint3",
            "bulletPoint4",
            "bulletPoint5",
        ],
        [
            "SKU-1",
            "Title One",
            "alpha beta",
            "<b>Description One</b>",
            "1. ONE - Content",
            "2. TWO - Content",
            "3. THREE - Content",
            "4. FOUR - Content",
            "5. FIVE - Content",
        ],
    )

    result = run_cli("validate-output", str(source), str(output))

    assert result.returncode == 2
    assert "Output headers do not match target schema" in result.stderr


def test_validate_output_rejects_extra_sheet(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "extra-sheet-output.xlsx"
    create_source_workbook(source)
    create_output_workbook_with_extra_sheet(
        output,
        TARGET_HEADERS,
        [
            "SKU-1",
            "Title One",
            "<b>Description One</b>",
            "alpha beta",
            "1. ONE - Content",
            "2. TWO - Content",
            "3. THREE - Content",
            "4. FOUR - Content",
            "5. FIVE - Content",
        ],
    )

    result = run_cli("validate-output", str(source), str(output))

    assert result.returncode == 2
    assert "only sheet allowed is listings" in result.stderr.lower()


def test_validate_output_rejects_empty_required_field(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "empty-field-output.xlsx"
    create_source_workbook(source)
    create_output_workbook(
        output,
        TARGET_HEADERS,
        [
            "SKU-1",
            "Title One",
            "",
            "alpha beta",
            "1. ONE - Content",
            "2. TWO - Content",
            "3. THREE - Content",
            "4. FOUR - Content",
            "5. FIVE - Content",
        ],
    )

    result = run_cli("validate-output", str(source), str(output))

    assert result.returncode == 2
    assert "empty required fields" in result.stderr
